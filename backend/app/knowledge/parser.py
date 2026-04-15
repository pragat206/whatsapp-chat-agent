"""Document parsing for knowledge ingestion.

Supports PDF, DOCX, TXT, CSV, XLSX, JSON.
Each parser returns plain text that will be chunked and embedded.
"""

import csv
import io
import json

from app.core.logging import get_logger

logger = get_logger("knowledge.parser")


def parse_pdf(file_bytes: bytes) -> str:
    """Extract text from PDF."""
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(file_bytes))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text.strip())
    return "\n\n".join(pages)


def parse_docx(file_bytes: bytes) -> str:
    """Extract text from DOCX."""
    from docx import Document

    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def parse_txt(file_bytes: bytes) -> str:
    """Plain text — decode and return."""
    return file_bytes.decode("utf-8", errors="replace")


def parse_csv(file_bytes: bytes) -> str:
    """Convert CSV rows to readable text."""
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    lines = []
    for row in reader:
        line = " | ".join(f"{k}: {v}" for k, v in row.items() if v)
        lines.append(line)
    return "\n".join(lines)


def parse_xlsx(file_bytes: bytes) -> str:
    """Extract text from Excel spreadsheets."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            values = [str(c) if c is not None else "" for c in row]
            if i == 0:
                headers = values
            else:
                line = " | ".join(
                    f"{headers[j] if j < len(headers) else f'col{j}'}: {v}"
                    for j, v in enumerate(values)
                    if v
                )
                if line:
                    lines.append(line)
    return "\n".join(lines)


def parse_json_file(file_bytes: bytes) -> str:
    """Convert JSON to readable text."""
    data = json.loads(file_bytes.decode("utf-8"))
    if isinstance(data, list):
        return "\n\n".join(json.dumps(item, indent=2) for item in data)
    return json.dumps(data, indent=2)


PARSERS = {
    "pdf": parse_pdf,
    "docx": parse_docx,
    "txt": parse_txt,
    "csv": parse_csv,
    "xlsx": parse_xlsx,
    "json": parse_json_file,
}


def parse_file(file_bytes: bytes, source_type: str) -> str:
    """Parse file content based on source type."""
    parser = PARSERS.get(source_type)
    if not parser:
        raise ValueError(f"Unsupported source type: {source_type}")
    return parser(file_bytes)
